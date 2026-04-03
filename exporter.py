# Copyright (c) 2026 Khushal Jain
# Licensed under the MIT License

import pandas as pd
import openpyxl
from data_engine import DataEngine

def export_to_excel(filepath: str) -> dict:
    try:
        # Load Data
        snapshots = DataEngine.get_all_snapshots()
        history = DataEngine.get_full_price_history()
        products = DataEngine.get_all_products()

        # Convert to DataFrames
        df_snap = pd.DataFrame(snapshots)
        df_hist = pd.DataFrame(history)
        df_prod = pd.DataFrame(products)

        # Write to Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_snap.to_excel(writer, sheet_name='Latest Snapshots', index=False)
            df_hist.to_excel(writer, sheet_name='Price History Log', index=False)
            df_prod.to_excel(writer, sheet_name='Product Master', index=False)
            
            # Format Worksheets
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Freeze the top header row
                worksheet.freeze_panes = worksheet['A2']
                
                # Auto-filter setup
                max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
                worksheet.auto_filter.ref = f"A1:{max_col}{worksheet.max_row}"
                
                # Apply Header formatting and Borders
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                
                # Auto-adjust column width
                for col in worksheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 3)
                    worksheet.column_dimensions[col_letter].width = adjusted_width

        return {"status": "success", "message": f"Successfully exported to {filepath}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
